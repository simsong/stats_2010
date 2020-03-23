#!/usr/bin/env python3
#
# geotree report tools
import datetime
import json
import os
import os.path
import re
import sqlite3
import sys
import time
import zipfile
import io
import logging
import numpy as np
import itertools
import statistics
import subprocess
import gc
import math
from collections import deque

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Protection, Font, Fill, Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.comments import Comment
import openpyxl.styles.colors as colors


thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

top_thick_border = Border( top=Side(border_style=BORDER_THICK, color='0000FF') )
top_border = Border( top=Side(border_style=BORDER_THIN, color='00000000') )
right_border = Border( right=Side(border_style=BORDER_THIN, color='00000000') )
right_thick_border = Border( right=Side(border_style=BORDER_THICK, color='00000000') )
bottom_border = Border( bottom=Side(border_style=BORDER_THIN, color='00000000') )
bottom_thick_border = Border( bottom=Side(border_style=BORDER_THICK, color='00007F') )

# https://htmlcolorcodes.com/

BOLD        = Font(bold=True)
CENTERED    = Alignment(horizontal='center')
YELLOW_FILL = PatternFill(fill_type='solid', start_color=colors.YELLOW, end_color=colors.YELLOW)
PINK_FILL   = PatternFill(fill_type='solid', start_color='ffb6c1', end_color='ffb6c1')
LIGHT_GREEN_FILL  = PatternFill(fill_type='solid', start_color='EAFAF1', end_color='EAF8F1')

#PR_FILL = PatternFill(fill_type='solid', start_color='F4D03F', end_color='F4D03F')       # yellow
PR_FILL     = PatternFill(fill_type='solid', start_color='F4D03F')       # yellow
STATE1_FILL = PatternFill(fill_type='solid', start_color='F2F4F4')   # silver
STATE2_FILL = PatternFill(fill_type='solid', start_color='A9CCE3') # blue

def darker(openpyxl_fill):
    rgb = openpyxl_fill.start_color.rgb
    (r,g,b) = [int(s,16) for s in [rgb[0:2],rgb[2:4],rgb[4:6]]]
    r = max(r-1,0)
    g = max(g-1,0)
    b = max(b-1,0)
    color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(fill_type='solid', start_color=color)
               
class RingBuffer:
    from collections import deque
    def __init__(self, data):
        self.data = deque(data)

    def append(self, x):
        self.data.append(x)

    def rotate(self):
        self.data.append( self.data.popleft())

    def next(self):
        value = self.data.popleft()
        self.data.append(value)
        return value

class EasyWorkbook(Workbook):
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)

    def clean(self):
        """Remove default 'Sheet' if exists"""
        if 'Sheet' in self:
            del self['Sheet']
        
    def format_rows(self,ws,*,min_row,max_row,column=None,skip_blank=True,min_col,max_col,fills=None,value_fills=None,stripe=False):
        """Apply colors to each row when column changes. If column is not specified, change every row."""
        if fills:
            fills = RingBuffer(fills)
            fill  = fills.next()
        prev_fill  = None
        prev_value = None
        make_darker = False
        for cellrow in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            column_value = cellrow[column-min_col if column is not None else 0].value
            if skip_blank and column_value=='': 
                continue

            make_darker  = not make_darker

            if not column_value:
                fill = None
            elif value_fills is not None and column_value in value_fills:
                fill = value_fills[column_value]
            elif column_value == prev_value:
                fill = prev_fill
                continue        #  don't change
            else:
                prev_fill   = fill       = fills.next()
                prev_value  = column_value
                make_darker = False

            #if make_darker and stripe:
            #    fill = darker(fill)

            for cell in cellrow:
                if fill:
                    cell.fill = fill

    def set_cell(ws,*,row,column,**kwargs):
        cell = ws.cell(row=row, column=column)
        for (key,value) in kwargs.items():
            setattr(cell, key, value)

# 
COL_PLEVEL = 1
COL_LEVEL_NAME  = 2
COL_NUM_LEVELS = 3
COL_SUBLEVEL_NAME = 4
COL_NUM_SUBLEVELS = 5
COL_MIN_FANOUT    = 6
COL_MIN_POPULATIONS = 17

def wb_setup_overview(ws):
    """Set up the root of the spreadsheet, adding notes and other information"""
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(COL_LEVEL_NAME)].width=20
    ws.column_dimensions[get_column_letter(COL_SUBLEVEL_NAME)].width=20
    ws.cell(row=2, column=COL_PLEVEL).value = 'Partition'
    ws.cell(row=2, column=COL_LEVEL_NAME).value = 'Level'
    ws.cell(row=2, column=COL_NUM_LEVELS).value = '# Levels'
    ws.cell(row=2, column=COL_SUBLEVEL_NAME).value = 'Sublevel'
    ws.cell(row=2, column=COL_NUM_SUBLEVELS).value = '# Sublevels'

    ws.cell(row=1, column=COL_MIN_FANOUT).value     = 'sublevel fanouts (interpolation=min)'
    ws.cell(row=1, column=COL_MIN_FANOUT).alignment = CENTERED
    ws.cell(row=1, column=COL_MIN_FANOUT).fill      = YELLOW_FILL
    ws.merge_cells(start_row=1, end_row=1, start_column=COL_MIN_FANOUT, end_column=COL_MIN_FANOUT+10)
    ws.cell(row=2, column=COL_MIN_FANOUT).value = 'min'
    for n in range(1,10):
        ws.cell(row=2, column=COL_MIN_FANOUT+n).value = f"{n*10}th pct."

    ws.cell(row=2, column=COL_MIN_FANOUT+10).value = 'max'
    for col in range(COL_MIN_FANOUT,COL_MIN_FANOUT+10+1):
        ws.column_dimensions[get_column_letter(col)].width=10
        ws.cell(row=2, column=col).alignment = CENTERED
        ws.cell(row=2, column=col).fill = YELLOW_FILL

    ws.cell(row=1, column=COL_MIN_POPULATIONS).value     = 'sublevel populations (interpolation=min)'
    ws.cell(row=1, column=COL_MIN_POPULATIONS).alignment = CENTERED
    ws.cell(row=1, column=COL_MIN_POPULATIONS).fill      = PINK_FILL
    ws.merge_cells(start_row=1, end_row=1, start_column=COL_MIN_POPULATIONS, end_column=COL_MIN_POPULATIONS+10)
    ws.cell(row=2, column=COL_MIN_POPULATIONS).value = 'min'
    for n in range(1,10):
        ws.cell(row=2, column=COL_MIN_POPULATIONS+n).value = f"{n*10}th pct."

    ws.cell(row=2, column=COL_MIN_POPULATIONS+10).value = 'max'
    for col in range(COL_MIN_POPULATIONS,COL_MIN_POPULATIONS+10+1):
        ws.column_dimensions[get_column_letter(col)].width=10
        ws.cell(row=2, column=col).alignment = CENTERED
        ws.cell(row=2, column=col).fill = PINK_FILL

    ws_bold_region(ws, min_row=1, max_row=2, min_col=1, max_col=COL_MIN_POPULATIONS+10)
    return 3                    # next row

WIDTH_9_DIGITS_COMMAS=12
COL_STUSAB    = 1
COL_PREFIX    = 2
COL_BYPASSED  = 3
COL_NAME      = 4
COL_FANOUT    = 5
COL_BLK_TOT   = 6
COL_POP_TOT   = 7
COL_POP_AVG   = 8
COL_POP_MIN   = 9

def ws_bold_region(ws,*,min_row,max_row,min_col,max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.font = BOLD
    

def ws_setup_level(ws, sheet_title, level):
    prefix = "Prefix " + ",".join([f"P{p}" for p in range(1,level+1)]) 

    ws.cell(row=2, column=COL_STUSAB).value = 'STUSAB'             # A2
    ws.cell(row=2, column=COL_PREFIX).value = prefix               # B2
    ws.column_dimensions[get_column_letter(COL_PREFIX)].width=25
    ws.cell(row=2, column=COL_BYPASSED).value = 'Bypassed'
    ws.cell(row=2, column=COL_NAME).value = 'Name'   # C2
    ws.column_dimensions[get_column_letter(COL_NAME)].width=20
    for col in range(COL_FANOUT,COL_POP_MIN+11+1):
        ws.column_dimensions[get_column_letter(col)].width=WIDTH_9_DIGITS_COMMAS

    ws.cell(row=1,column=COL_FANOUT).value = sheet_title
    ws.cell(row=1,column=COL_FANOUT).alignment = CENTERED
    ws.merge_cells(start_row=1, end_row=1, start_column=COL_FANOUT, end_column=COL_POP_MIN+10)
    ws.cell(row=1,column=COL_POP_AVG+9).border = right_border
    ws.cell(row=2,column=COL_FANOUT).value='fanout'

    ws.cell(row=2,column=COL_BLK_TOT).value='# blocks'
    ws.cell(row=2,column=COL_POP_TOT).value='pop_tot'
    ws.cell(row=2,column=COL_POP_AVG).value='pop_avg'
    ws.cell(row=2,column=COL_POP_MIN).value='min pop'
    for n in range(1,10):
        ws.cell(row=2,column=COL_POP_MIN+n).value = f"{n*10}th pct."
    ws.cell(row=2,column=18).value='max pop'
    for col in range(COL_FANOUT,COL_POP_MIN+10+1):
        ws.cell(row=2,column=col).alignment = CENTERED
    ws.cell(row=2,column=COL_FANOUT).border = right_border
    ws.cell(row=2,column=COL_POP_AVG).border = right_thick_border
    ws.cell(row=2,column=COL_POP_MIN+10).border = right_border

    ws_bold_region(ws, min_row=2, max_row=2, min_col=1, max_col=COL_POP_MIN+10)

def ws_add_metadata(wb):
    ws = wb.create_sheet("Notes")
    ws.cell(row=1, column=1).value = "Command Line"
    ws.cell(row=1, column=2).value = " ".join([sys.executable] + sys.argv)

def ws_add_notes(ws,*,row,column,text):
    for line in text.split("\n"):
        if line[0:1]=='=':
            line = ' ' + line
        ws.cell(row=row, column=column).value = line.rstrip()
        row += 1
            

