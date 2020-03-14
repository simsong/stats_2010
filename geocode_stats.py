#!/usr/bin/env python3
"""
Goals of this program:

- We are interested in building a tree of geounits. For each geounit, we want to measure fanout. 

- We want to be able to experiment with different partitioning functions, and to explore partitioning functions that are based on the data

- We want to consider all geographies, inhabited 2010 geographies, and 2010 geographies with housing units or GQ facilities.


https://openpyxl.readthedocs.io/en/stable/filters.html - add filters
"""

import os
import sys
import time
import re
import sqlite3
import statistics
import numpy as np
import constants
import ctools.clogging
import logging
import subprocess
import itertools
from collections import deque

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Protection, Font, Fill, Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.comments import Comment
import openpyxl.styles.colors as colors

import pl94_dbload

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

right_border = Border(
    right=Side(border_style=BORDER_THIN, color='00000000')
)
BOLD     = Font(bold=True)
CENTERED = Alignment(horizontal='center')
YELLOW_FILL   = PatternFill(fill_type='solid', start_color=colors.YELLOW, end_color=colors.YELLOW)
PINK_FILL   = PatternFill(fill_type='solid', start_color='ffb6c1', end_color='ffb6c1')

def deciles(ary):
    return np.percentile(ary, np.arange(0,101,10), interpolation='lower')

def flatmap(func, *iterable):
    return itertools.chain.from_iterable(map(func, *iterable))

class GeocodeStats:
    def __init__(self,args):
        self.args = args
        self.conn = sqlite3.connect(args.db)
        self.conn.row_factory = sqlite3.Row # give me dicts
        if self.args.geocode2:
            self.geocode = 'geocode2'
        elif self.args.geocode3:
            self.geocode = 'geocode3'
        else:
            self.geocode = 'geocode'

        c = self.cursor()
        c.execute(f"PRAGMA cache_size = {-1024*1024*16}")  # 1GiB cache
        c.execute("SELECT state,county,name FROM geo where sumlev=50") 
        self.counties = {(row['state'],row['county']):row['name'] for row in c.fetchall()}

    def cursor(self):
        return self.conn.cursor()

    def county_name(self,state,county):
        return self.counties[(state,county)]

    def group_blocks_by_prefix(self,prefix_len):
        """Group all of the blocks by a given prefix and return aggregate statistics. 
        They then get re-arregated at a higher level by the caller"""
        grouper  = f"SUBSTR({self.geocode},1,{prefix_len}) "
        c = self.cursor()
        cmd = f"""SELECT {grouper} AS prefix, state, county, aiannh,
                  SUM(1) AS block_count,
                  SUM( CASE WHEN pop>0 THEN 1 ELSE 0 END) as populated_block_count,
                  SUM(pop) AS population
                  FROM blocks """
        cmd += f"""GROUP BY {grouper} ORDER BY 1 """
        logging.info(re.sub(' +',' ',cmd.replace("\n"," ")))
        t0 = time.time()
        c.execute(cmd)
        res = c.fetchall()
        t1 = time.time()
        logging.info("QUERY TIME: %s  QUERY RESULTS: %s RESULTS/SEC: %s",
                     t1-t0,len(res),len(res)/(t1-t0))
        return res

    def geocode3_name(self,gc):
        """Decode the GEOCODE3 name"""
        append = ""
        if len(gc)>=11:
            # PLACE
            if gc[0:6]=='11XXXX':
                query=f"SELECT name FROM GEO WHERE state=11        AND '{gc[8:11]}'=sldu AND sumlev=610" # DC wards
            elif gc[2]=='X' and gc[3].isdigit():
                # not strong MCD states
                query=f"SELECT name FROM GEO WHERE {gc[0:2]}=state AND {gc[6:11]}=place AND sumlev=160"
            elif gc[2]=='X':
                # Strong MCD states
                query=f"SELECT name FROM GEO WHERE {gc[0:2]}=state AND {gc[6:11]}=cousub AND sumlev=521"
            else:
                # AIANNH
                query=f"SELECT name FROM GEO WHERE {gc[0:2]}=state AND {gc[6:11]}=aiannh AND sumlev=280"
        elif len(gc)>=6:       
            # COUNTY
            if gc[0:6]=='11XXXX':
                query=f"SELECT 'Washington DC'"
            elif gc[2]=='X' and gc[3].isdigit():
                # not strong MCD states
                query=f"SELECT NAME from geo WHERE state={gc[0:2]} AND county={gc[3:6]}"
            else:
                query=f"SELECT 'COUNTY IGNORED'"
        elif len(gc)>=3:       
            # STATE
            query=f"SELECT name FROM GEO WHERE state={gc[0:2]} AND sumlev=40"
            if gc[2]=='A':
                append=" (AIAN AREAS)"
        else:
            query="SELECT 'USA'"
        query += " LIMIT 1"
        c = self.cursor()
        try:
            c.execute(query)
        except sqlite3.OperationalError as e:
            print(query,file=sys.stderr)
            print(e)
            assert(0)
        res = c.fetchone()
        if res is None:
            if gc[6:11]=='99999' or gc[6:11]=='00000':
                return "(none)"
            print(f"Query for {gc} returned no values: {query}",file=sys.stderr)
            return ""
        res = res[0]
        if append:
            res += " " + append
        if len(gc) > 14:
            res += "TRACT " + gc[14:20]
        if len(gc) > 20:
            res += "BLKGRP " + gc[20:22]
        if len(gc) > 22:
            res += "BLOCK " + gc[22:26]
        return res

    def fanout_name(self,fanout_len):
        if self.args.geocode2:
            ary = [ ('state',2), ('county',3), ('cousub',5), ('tract',4), ('block',4)]
        elif self.args.geocode3:
            ary = [ ('AIAN state',3), ('county',3), ('place',5), ('tract',9), ('blkgrp',2), ('block',4)]
        else:
            ary = [ ('state',2), ('county',3), ('tract',4), ('block',4)]
            
        msg = 'Fanout to '
        try:
            while fanout_len > 0:
                s = min(fanout_len,ary[0][1])
                msg += f"[{ary[0][0]}:{s}]"
                fanout_len -= s
                ary.pop(0)
        except IndexError:
            msg + f"... {fanout_len}"
        return msg

    def add_geocode_report(self,wb,report_len,fanout_len):
        """Create a report for a given report_len and child_len. A span=2 is by state, a span=5 is state and county."""
        # First generate the data
        data         = deque(gs.group_blocks_by_prefix(fanout_len))
        total_population = 0

        # Now find all of the fanout groups. Each fanout group will become a line in the report.
        fanout_groups = []
        logging.info(f"report_len: {report_len} len(data): {len(data)}")
        while data:
            details = []
            res = {}
            res['reporting_prefix'] = data[0]['prefix'][0:report_len]
            res['state']  = state  = data[0]['state']
            res['county'] = county = data[0]['county']
            populations = []
            while data and (res['reporting_prefix'] == data[0]['prefix'][0:report_len]):
                d0 = data.popleft()
                if args.details:
                    details.append((d0['prefix'],d0['population']))
                group_population = d0['population']
                total_population += group_population
                populations.append( group_population )
            res['populations'] = populations
            res['details'] = details
            fanout_groups.append(res)
        logging.info("len(fanout_groups) = %s",len(fanout_groups))
        logging.info("Data have been computed. Now laying out spreadsheet")
        ws = wb.create_sheet(f'S {fanout_len}')
        if 'Sheet' in wb:
            del wb['Sheet']
        ws.cell(row=2, column=1).value = 'STUSAB' # A2
        ws.cell(row=2, column=2).value = 'Prefix' # B2
        ws.cell(row=2, column=3).value = 'Name'   # C2
        ws.column_dimensions['C'].width=20
        for ch in 'DEFGHIJKLMNOPQ':
            ws.column_dimensions[ch].width=10

        ws.cell(row=1,column=4).value = self.fanout_name(fanout_len)
        ws.cell(row=1,column=4).alignment = CENTERED
        ws.merge_cells(start_row=1,end_row=1,start_column=4,end_column=7+10)
        ws.cell(row=1,column=6+9).border = right_border
        ws.cell(row=2,column=4).value='fanout'
        ws.cell(row=2,column=5).value='pop_tot'
        ws.cell(row=2,column=6).value='pop_avg'
        ws.cell(row=2,column=7).value='min pop'
        for n in range(1,10):
            ws.cell(row=2,column=7+n).value = f"{n*10}th pct."
        ws.cell(row=2,column=17).value='max pop'
        for col in range(4,7+10+1):
            ws.cell(row=2,column=col).alignment = CENTERED
        ws.cell(row=2,column=4).border = right_border
        ws.cell(row=2,column=6).border = right_border
        ws.cell(row=2,column=7+10).border = right_border

        # Using the grouper that we have specified for children, get the report
        row = 3
        for res in fanout_groups:
            if row%10_000==0:
                logging.info("row=%s",row)
            state  = res['state']
            county = res['county']
            populations = res['populations']
            reporting_prefix = res['reporting_prefix']

            ws.cell(row=row,column=1).value = constants.STATE_TO_STUSAB[state]
            ws.cell(row=row,column=2).value = res['reporting_prefix']
            if args.names:
                if self.args.geocode3 and len(reporting_prefix)<12:
                    ws.cell(row=row,column=3).value = gs.geocode3_name(reporting_prefix)
                else:
                    ws.cell(row=row,column=3).value = gs.county_name(state, county)
            ws.cell( row=row,column=4).value = len(populations)
            ws.cell( row=row,column=5).value = sum(populations)
            ws.cell( row=row,column=6).value = int(statistics.mean(populations))
            for cellrow in ws.iter_rows(min_row=row, max_row=row,min_col=7,max_col=17):
                for(cell,value) in zip(cellrow, deciles(populations)):
                    cell.value = value
            row += 1

        logging.info("now adding borders")
        for cellrow in ws.iter_rows(min_row=2, max_row=row-1, min_col=1, max_col=17):
            cellrow[3].border = right_border
            cellrow[5].border = right_border
            cellrow[16].border = right_border

        logging.info("now freezing")
        # Increase Spreadsheet usability
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:Q{row-1}'
        row += 1

        ws.cell( row=row,column=3).value = "Total population:"
        ws.cell( row=row,column=5).value = total_population
        row += 2

        if args.details:
            # Provide details
            logging.info("adding details")
            ws.cell(column=3, row=row).value = "Details:"
            row += 1
            for res in fanout_groups:
                cell       = ws.cell(column=3, row=row)
                cell.value = res['reporting_prefix']
                cell.font  = BOLD
                row += 1
                for (a,b) in res['details']:
                    ws.cell( row=row,column=3).value=a
                    ws.cell( row=row,column=4).value=b
                    row += 1
                row += 1

    def geolevel_report_setup(self,ws):
        ws.cell(row=2, column=1).value = 'Level'
        ws.cell(row=2, column=2).value = '# Levels'
        ws.cell(row=2, column=3).value = 'Sublevel'
        ws.cell(row=2, column=4).value = '# Sublevels'

        ws.cell(row=1, column=5).value     = 'sublevel fanouts (interpolation=min)'
        ws.cell(row=1, column=5).alignment = CENTERED
        ws.cell(row=1, column=5).fill      = YELLOW_FILL
        ws.merge_cells(start_row=1, end_row=1, start_column=5, end_column=15)
        ws.cell(row=2, column=5).value = 'min'
        for n in range(1,10):
            ws.cell(row=2, column=5+n).value = f"{n*10}th pct."
        ws.cell(row=2, column=15).value = 'max'
        for col in range(5,16):
            ws.cell(row=2, column=col).alignment = CENTERED
            ws.cell(row=2, column=col).fill = YELLOW_FILL
        
        ws.cell(row=1, column=16).value     = 'sublevel populations (interpolation=min)'
        ws.cell(row=1, column=16).alignment = CENTERED
        ws.cell(row=1, column=16).fill      = PINK_FILL
        ws.merge_cells(start_row=1, end_row=1, start_column=16, end_column=26)
        ws.cell(row=2, column=16).value = 'min'
        for n in range(1,10):
            ws.cell(row=2, column=16+n).value = f"{n*10}th pct."
        ws.cell(row=2, column=26).value = 'max'
        for col in range(16,27):
            ws.cell(row=2, column=col).alignment = CENTERED
            ws.cell(row=2, column=col).fill = PINK_FILL
        
    def add_geolevel_report(self,ws,ct,level_name,sublevel_name,level_len,sublevel_len):
        """Create a report of the fanout for a given geolevel."""

        # First generate the data
        row  = ct+4
        data = deque(gs.group_blocks_by_prefix(sublevel_len))

        # Now find all of the fanout for this level
        logging.info(f"report_len: {level_name} len(data): {len(data)}")
        level_stats = []
        while data:
            res = {}
            level = data[0]['prefix'][0:level_len]
            sublevel_populations = []
            while data and (level == data[0]['prefix'][0:level_len]):
                d0 = data.popleft()
                sublevel_populations.append(d0['population'])
            res['sublevel_count']       = len(sublevel_populations)
            res['sublevel_populations'] = sublevel_populations
            level_stats.append(res)
            if len(level_stats)%100000==0:
                logging.info("len(level_stats)=%d",len(level_stats))
        logging.info("len(level_stats) = %s",len(level_stats))
        ws.cell(row=row, column=1).value = level_name
        ws.cell(row=row, column=2).value = len(level_stats)
        ws.cell(row=row, column=3).value = sublevel_name
        ws.cell(row=row, column=4).value = sum([res['sublevel_count'] for res in level_stats])

        sublevel_fanouts = [res['sublevel_count'] for res in level_stats]
        sublevel_fanout_deciles = deciles(sublevel_fanouts)
        for cellrow in ws.iter_rows(min_row = row, max_row=row, min_col=5, max_col = 15):
            for (cell,value) in zip(cellrow,sublevel_fanout_deciles):
                cell.value = value
                cell.fill = YELLOW_FILL

        all_sublevel_populations = [res['sublevel_populations'] for res in level_stats]
        all_sublevel_populations = list(flatmap( lambda a:a, all_sublevel_populations))
        sublevel_population_deciles = deciles(all_sublevel_populations)
        for cellrow in ws.iter_rows(min_row = row, max_row=row, min_col=16, max_col = 26):
            for (cell,value) in zip(cellrow,sublevel_population_deciles):
                cell.value = value
                cell.fill = PINK_FILL



if __name__=="__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Report statistics about geocode prefixes',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--prefix", help="Report about a prefix")
    parser.add_argument("--fanout_len", type=int, help="Specify number of characters for fanout_len")
    parser.add_argument("--db", help="Specify database location", default=pl94_dbload.DBFILE)
    parser.add_argument("--limit", type=int, help="Return only this many")
    parser.add_argument("--geocode2", action='store_true', help="Use geocode2, not geocode")
    parser.add_argument("--geocode3", action='store_true', help="Use geocode3, not geocode")
    parser.add_argument("--geocode_report", help="Generate a fanout report by geocode", action='store_true')
    parser.add_argument("--geolevel_report", help="Generate a fanout report by geolevel", action='store_true')
    parser.add_argument("--prefixset", help="specify the sets of prefixes to use, in the form prefix:child_len,prefix:child_len")
    parser.add_argument("--debug", action='store_true')
    parser.add_argument("--details", action='store_true', help='Provide details for each fanout')
    parser.add_argument("--open" , action='store_true', help='automatically open on finish')
    parser.add_argument("--names", action='store_true', help='display names')
    ctools.clogging.add_argument(parser)
    args = parser.parse_args()
    ctools.clogging.setup(level = args.loglevel)
    gs = GeocodeStats(args)
    wb = Workbook()

    if args.geocode_report:
        vintage = "report geocode " + time.strftime("%Y-%m-%d %H%M%S")
        for (ct,report_desc) in enumerate(args.prefixset.split(",")):
            (prefix_len,fanout_len) = [int(x) for x in report_desc.split(":")]
            gs.add_geocode_report(wb,prefix_len,fanout_len)
            fname = f"{vintage} {ct:02}.xlsx"
            logging.info("Saving %s",fname)
            wb.save(fname)
            
    if args.geolevel_report:
        ws = wb.active
        vintage = "report geolevel " + time.strftime("%Y-%m-%d %H%M%S")
        gs.geolevel_report_setup(ws)
        for (ct,report_desc) in enumerate(args.prefixset.split(",")):
            logging.info(report_desc)
            (level_name,sublevel_name,level_len,sublevel_len) = report_desc.split(":")
            level_len = int(level_len)
            sublevel_len  = int(sublevel_len)
            gs.add_geolevel_report(ws, ct, level_name, sublevel_name, level_len, sublevel_len)
            fname = f"{vintage} {ct:02}.xlsx"
            logging.info("Saving %s",fname)
            wb.save(fname)
    if args.open:
        subprocess.call(['open',fname])

