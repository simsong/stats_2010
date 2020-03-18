__version__ = '0.2.0'
import datetime
import json
import os
import os.path
import re
import sqlite3
import sys
import time
import zipfile
import io
import logging
import numpy as np
import itertools
import statistics
import subprocess
import gc
import math
from collections import deque

#                       P0        P1                  P2                  P3          P4         P5      P6
GEOTREE={'v1':{'names':['US',    'DC•STATE',          'COUNTY',           'TGROUP',   'TRACT',   'BGROUP','BLOCK'],
               'name':'Geography used for 2010 Demonstration Data Products' },

         'v2':{'names':['US•PR' ,'DC•STATE•ASTATE•PR','SLDU•COUNTY•PLACE','TRACT',    'BLKGRP2', 'BLOCK', None],
               'name':'Revised MCD and AIAN-aware geography' },

         'v3':{'names':['US•PR' ,'DC•STATE•ASTATE•PR','SLDU•COUNTY•PLACE','LEVEL3',   'BLOCK',  None, None],
               'name':'Revised MCD and AIAN-aware geography with synthetic LEVEL3' 
               }
         }



from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Protection, Font, Fill, Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.comments import Comment
import openpyxl.styles.colors as colors


thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

top_thick_border = Border( top=Side(border_style=BORDER_THICK, color='0000FF') )
top_border = Border( top=Side(border_style=BORDER_THIN, color='00000000') )
right_border = Border( right=Side(border_style=BORDER_THIN, color='00000000') )
right_thick_border = Border( right=Side(border_style=BORDER_THICK, color='00000000') )
bottom_border = Border( bottom=Side(border_style=BORDER_THIN, color='00000000') )
bottom_thick_border = Border( bottom=Side(border_style=BORDER_THICK, color='00007F') )

# https://htmlcolorcodes.com/

BOLD     = Font(bold=True)
CENTERED = Alignment(horizontal='center')
YELLOW_FILL   = PatternFill(fill_type='solid', start_color=colors.YELLOW, end_color=colors.YELLOW)
PINK_FILL   = PatternFill(fill_type='solid', start_color='ffb6c1', end_color='ffb6c1')
LIGHT_GREEN_FILL  = PatternFill(fill_type='solid', start_color='EAFAF1', end_color='EAF8F1')

#PR_FILL = PatternFill(fill_type='solid', start_color='F4D03F', end_color='F4D03F')       # yellow
PR_FILL     = PatternFill(fill_type='solid', start_color='F4D03F')       # yellow
STATE1_FILL = PatternFill(fill_type='solid', start_color='F2F4F4')   # silver
STATE2_FILL = PatternFill(fill_type='solid', start_color='A9CCE3') # blue

def darker(openpyxl_fill):
    rgb = openpyxl_fill.start_color.rgb
    (r,g,b) = [int(s,16) for s in [rgb[0:2],rgb[2:4],rgb[4:6]]]
    r = max(r-1,0)
    g = max(g-1,0)
    b = max(b-1,0)
    color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(fill_type='solid', start_color=color)
               

import pl94_dbload
import ctools.dbfile
import ctools.clogging
import ctools.timer
import constants
from constants import *

def deciles(ary):
    return np.percentile(ary, np.arange(0,101,10), interpolation='lower')

def flatmap(func, *iterable):
    return itertools.chain.from_iterable(map(func, *iterable))

###
### Create geotree schema
###
CREATE_GEOTREE_SQL="""
CREATE TABLE %TABLE% (state INTEGER,
                      logrecno INTEGER,
                      p1 VARCHAR(16),
                      p2 VARCHAR(16),
                      p3 VARCHAR(16),
                      p4 VARCHAR(16),
                      p5 VARCHAR(16),
                      p6 VARCHAR(16));
CREATE UNIQUE INDEX %TABLE%_logrecno ON %TABLE%(state,logrecno);
CREATE UNIQUE INDEX %TABLE%_p ON %TABLE%(p1,p2,p3,p4,p5,p6);
"""


class RingBuffer:
    from collections import deque
    def __init__(self, data):
        self.data = deque(data)

    def append(self, x):
        self.data.append(x)

    def rotate(self):
        self.data.append( self.data.popleft())

    def next(self):
        value = self.data.popleft()
        self.data.append(value)
        return value

class EasyWorkbook(Workbook):
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)

    def clean(self):
        """Remove default 'Sheet' if exists"""
        if 'Sheet' in self:
            del self['Sheet']
        
    def format_rows(self,ws,*,min_row,max_row,column=None,skip_blank=True,min_col,max_col,fills=None,value_fills=None,stripe=False):
        """Apply colors to each row when column changes. If column is not specified, change every row."""
        if fills:
            fills = RingBuffer(fills)
            fill  = fills.next()
        prev_value = None
        make_darker = False
        for cellrow in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            column_value = cellrow[column-min_col if column is not None else 0].value
            if skip_blank and column_value=='': 
                continue

            make_darker  = not make_darker

            if (value_fills is not None) and (column_value in value_fills):
                fill = value_fills[column_value]
            elif fills:
                if column_value != prev_value:
                    fill = fills.next()
                    prev_value = column_value
                    make_darker = False
            else:
                fill = None

            #if make_darker and stripe:
            #    fill = darker(fill)

            for cell in cellrow:
                if fill:
                    cell.fill = fill

    def set_cell(ws,*,row,column,**kwargs):
        cell = ws.cell(row=row, column=column)
        for (key,value) in kwargs.items():
            setattr(cell, key, value)

    
# 
def wb_setup_overview(ws):
    """Set up the root of the spreadsheet, adding notes and other information"""
    ws.cell(row=2, column=1).value = 'Level'
    ws.cell(row=2, column=2).value = '# Levels'
    ws.cell(row=2, column=3).value = 'Sublevel'
    ws.cell(row=2, column=4).value = '# Sublevels'

    ws.cell(row=1, column=5).value     = 'sublevel fanouts (interpolation=min)'
    ws.cell(row=1, column=5).alignment = CENTERED
    ws.cell(row=1, column=5).fill      = YELLOW_FILL
    ws.merge_cells(start_row=1, end_row=1, start_column=5, end_column=15)
    ws.cell(row=2, column=5).value = 'min'
    for n in range(1,10):
        ws.cell(row=2, column=5+n).value = f"{n*10}th pct."
    ws.cell(row=2, column=15).value = 'max'
    for col in range(5,16):
        ws.cell(row=2, column=col).alignment = CENTERED
        ws.cell(row=2, column=col).fill = YELLOW_FILL

    ws.cell(row=1, column=16).value     = 'sublevel populations (interpolation=min)'
    ws.cell(row=1, column=16).alignment = CENTERED
    ws.cell(row=1, column=16).fill      = PINK_FILL
    ws.merge_cells(start_row=1, end_row=1, start_column=16, end_column=26)
    ws.cell(row=2, column=16).value = 'min'
    for n in range(1,10):
        ws.cell(row=2, column=16+n).value = f"{n*10}th pct."
    ws.cell(row=2, column=26).value = 'max'
    for col in range(16,27):
        ws.cell(row=2, column=col).alignment = CENTERED
        ws.cell(row=2, column=col).fill = PINK_FILL
    return 3                    # next row

def ws_setup_level(ws,sheet_title):
    ws.cell(row=2, column=1).value = 'STUSAB' # A2
    ws.cell(row=2, column=2).value = 'Prefix' # B2
    ws.cell(row=2, column=3).value = 'Name'   # C2
    ws.column_dimensions['C'].width=20
    for ch in 'DEFGHIJKLMNOPQ':
        ws.column_dimensions[ch].width=10

    ws.cell(row=1,column=4).value = sheet_title
    ws.cell(row=1,column=4).alignment = CENTERED
    ws.merge_cells(start_row=1,end_row=1,start_column=4,end_column=7+10)
    ws.cell(row=1,column=6+9).border = right_border
    ws.cell(row=2,column=4).value='fanout'
    ws.cell(row=2,column=5).value='pop_tot'
    ws.cell(row=2,column=6).value='pop_avg'
    ws.cell(row=2,column=7).value='min pop'
    for n in range(1,10):
        ws.cell(row=2,column=7+n).value = f"{n*10}th pct."
    ws.cell(row=2,column=17).value='max pop'
    for col in range(4,7+10+1):
        ws.cell(row=2,column=col).alignment = CENTERED
    ws.cell(row=2,column=4).border = right_border
    ws.cell(row=2,column=6).border = right_thick_border
    ws.cell(row=2,column=7+10).border = right_border

def ws_add_notes(ws,row,fname):
    ws.cell(row=row, column=1).value = "Notes"
    row += 1
    with open(fname,"r") as f:
        for line in f:
            ws.cell(row=row, column=1).value = line.strip()
            row += 1
            
def ws_add_metadata(wb):
    ws = wb.create_sheet("Notes")
    ws.cell(row=1, column=1).value = "Command Line"
    ws.cell(row=1, column=2).value = " ".join([sys.executable] + sys.argv)


def include_aianhh(code):
    if 1 <= code <= 4999:
        return "Federally recognized American Indian Reservations and Off-Reservation Trust Lands"
    elif 5000 <= code  <=5999:
        return "Hawaiian Home Lands"
    elif 6000 <= code <= 7999:
        return "Alaska Native Village Statistical Areas"
    elif 9000 <= code <= 9499:
        if EXCLUDE_STATE_RECOGNIZED_TRIBES:
            return False
        else:
            return "State recognized American Indian Reservations"
    else:
        return False

NE_STATES="ME,NH,VT,MA,CT,RI"
V2_STRONG_MCD_STATES=[9,11,23,25,26,27,33,34,36,42,44,50,55]
V21_STATE_LIST=[STUSAB_TO_STATE[stusab] for stusab in ",".split("MI,MN,NJ,NY,PA,WI" + "," + NEW_ENGLAND_STATES)]
EXCLUDE_STATE_RECOGNIZED_TRIBES=True
def geocode3(gh,scheme='v2'):
    """The revised geocode that takes into account AIANHH. Levels are:
    0 - US or PR
    1 - Non-AIANHH part-of-state or PR            | AIANHH part-of-State 
    2 - COUNTY in non-strong MCD states           | ignored in AIANHH
    3 - PLACE in 38 strong-MCD states, SLDU in DC | AIANHH in AIANHH states
    4 - TRACT or 3-digit TG or 4-digit TG         | TRACT
    5 - BLKGRP first 1 or 2 digits of block       | BLKGRP
    6 - BLOCK                                     | BLOCK
    """
    block  = f"{gh['block']:04}"
    blkgrp2 = block[0:2]         # note 2-digit block groups
    if gh['state']==DC_FIPS:
        # Washington DC
        return (f"{DC_FIPS:02}D",    f"____{int(gh['sldu']):05}",            f"___{gh['tract']:06}",               blkgrp2, block, None )
    if gh['state']==PR_FIPS:
        # Puerto Rico
        return (f"{PR_FIPS:02}P",    f"{gh['county']:03}{gh['place']:05}",   f"___{gh['tract']:06}",               blkgrp2, block, None )
    elif include_aianhh(gh['AIANHH']):
        # AIANHH portion of 38-states with AIANHH
        return (f"{gh['state']:02}A", f"{gh['aianhh']:05}{gh['county']:03}", f"___{gh['tract']:06}",               blkgrp2, block, None )
    elif (gh['STATE'] in V21_STATE_LIST and scheme=='v2.1'):
        # Non-AIAN area in 12 states that have so many strong MCDs that we need to group them by county
        return (f"{gh['state']:02}X", f"{gh['tabblkst']:05}{gh['tabblkcou']}{gh['cousubfp']:05}",               f"{gh['county']:03}{gh['tract']:06}", blkgrp2, block, None  )
    elif gh['STATE'] in V2_STRONG_MCD_STATES:
        # Non-AIAN area in 12 state with strong MCD.
        # County is included in tract to make it unique, but cousubs do not cross counties.
        return (f"{gh['state']:02}X", f"___{gh['cousub']:05}",               f"{gh['county']:03}{gh['tract']:06}", blkgrp2, block, None  )
    else:
        # Non-AIAN area and 38 states not strong MCD
        return (f"{gh['state']:02}X", f"{gh['county']:03}{gh['place']:05}",  f"___{gh['tract']:06}",               blkgrp2, block, None  )
    return "".join(code)


class MinMax:
    """Remember an object associated with a min and the object associated with the max."""
    def __int__(self,func):
        self.func = func
        self.the_min  = None
        self.the_max  = None
    def add(self,obj):
        val = func(obj)

class GeoTree:
    def __init__(self,db,name,scheme,xpr):
        self.db   = db
        self.name = name        # which table we are using
        self.scheme = scheme
        self.gt     = GEOTREE[scheme]
        self.xpr    = xpr

    def create(self):
        if self.scheme=='v1':
            self.db.create_schema(CREATE_GEOTREE_SQL.replace("%TABLE%",self.name),debug=True)
            cmd = f"""INSERT INTO {self.name} 
            SELECT state AS state,
            logrecno AS logrecno,
            printf("%02d",state) as p1,
            printf("%03d",county) as p2,
            substr(printf("%05d",tract),1,2) as p3,
            printf("%06d",tract) as p4,
            substr(printf("%04d",block),1,1) as p5,
            printf("%04d",block) as p6
            FROM blocks"""
            self.db.execute(cmd,debug=True)
            self.db.commit()
        elif self.scheme=='v2':
            # This could be made a LOT faster. Right now we just go row-by-row
            # It takes about 5 minutes.
            self.db.create_schema(CREATE_GEOTREE_SQL.replace("%TABLE%",self.name),debug=True)
            c = self.db.execute("SELECT * from blocks")
            for block in c:
                try:
                    p = geocode3(block)
                except KeyError as e:
                    print("block:",block,file=sys.stderr)
                    print(e,file=sys.stderr)
                    exit(1)
                self.db.execute(f"INSERT INTO {self.name} (state,logrecno,p1,p2,p3,p4,p5,p6) values (?,?,?,?,?,?,?,?)",
                                (block['STATE'],block['LOGRECNO'],p[0],p[1],p[2],p[3],p[4],p[5]))
            self.db.commit()
        elif self.scheme=='v3':
            # V2 is the v2 geography for p1 and P2, but an adaptive algorithm for P3 and P4. There is no P5.
            # 
            self.db.create_schema(CREATE_GEOTREE_SQL.replace("%TABLE%",self.name),debug=True)
            c = self.db.execute(f"SELECT state,p1,p2,count(*) as count from table2 group by state,p1,p2")
            for row in c:
                state   = row['state']
                fanout1 = int(math.sqrt(row['count']))
                p1 = row['p1']
                p2 = row['p2']
                p3 = p4 = 1
                d = self.db.execute(f"SELECT logrecno from table2 where state=? and p1=? and p2=?",
                                    (row['state'],p1,p2))
                for row2 in d:
                    self.db.execute(f"INSERT into {self.name} (state,logrecno,p1,p2,p3,p4) values (?,?,?,?,?,?)",
                                    (state,row2['logrecno'],p1,p2,format(p3,"04"),format(p4,"04")))
                    p4 += 1
                    if p4 >= fanout1:
                        p3 += 1
                        p4 = 1
                logging.info("Completed %s %s %s",state,p1,p2)
            self.db.commit()
                    

        else:
            raise RuntimeError(f"Unknown scheme: {self.scheme}")

    def dump(self):
        cmd = f"""select a.state,a.logrecno,a.p1,a.p2,a.p3,a.p4,a.p5,a.p6,b.geocode,b.pop 
        from {self.name} a left join blocks b on a.state=b.state and a.logrecno=b.logrecno"""
        c = self.db.execute(cmd)
        for row in c:
            print(",".join([str(x) for x in row]))

    def get_geounits(self,ct):
        """ 
        When ct=0, nation page is constructed, single nation row goes to subpops.
        When ct=1, state page is being constructed, state rows need to be counties"""
        reporting_prefix = "||' '||".join(["''"] + [f"p{n}" for n in range(1,ct+1)])
        plevel1 = ",".join([f"p{n}" for n in range(1,ct+1)])  # if ct=0, this is P1
        plevel2 = ",".join([f"p{n}" for n in range(1,ct+2)])  # if ct=0, this is P1,P2
        plevel3 = ",".join([f"p{n}" for n in range(1,ct+3)])  # if ct=0, this is P1,P2,P3
        plevel4 = ",".join([f"p{n}" for n in range(1,ct+4)])  # if ct=0, this is P1,P2,P3,P4
        logging.info(f"ct:{ct} plevel1:{plevel1}")

        if self.xpr:
            where = f'WHERE a.state!={PR_FIPS} and b.state!={PR_FIPS}'
        else:
            where = ''
        cmd = f"""SELECT a.state,{reporting_prefix} as reporting_prefix,{plevel2},COUNT(*) as count,SUM(pop) as population FROM 
        {self.name} a LEFT JOIN blocks b ON a.state=b.state AND a.logrecno=b.logrecno {where} GROUP BY {plevel2}"""
        c = self.db.execute(cmd)
        t0 = time.time()
        res = c.fetchall()
        t1 = time.time()
        logging.info(f"Query Time: {t1-t0}")
        return res

    def end_state(self,ws,start_row,end_row):
        if start_row + 1 > end_row:
            return
        for cellrow in ws.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=17):
            for cell in cellrow:
                cell.border = top_thick_border
        #  don't do grouping; it is confusing
        #ws.row_dimensions.group(start_row,end_row-1,hidden=True)

    def report(self):
        """Generate a geotree report into a spreadsheet. 
        Sheet overview      - report of all summary levels
        Sheet FANLEV    - report of fanout to that level.
        When ct=0, we are doing NATION on the overview and STATES on the tab.
        """

        vintage   = time.strftime("%Y-%m-%d %H%M%S")
        fnamebase = f"reports/{args.scheme} report-{vintage}"
        wb = EasyWorkbook()
        ws_overview = wb.create_sheet("Overview")
        wb.clean()
        overview_row = wb_setup_overview(ws_overview)
        for (ct,overview_name) in enumerate(self.gt['names'][0:-1]):
            t0 = time.time()
            fanout_name     = self.gt['names'][ct]
            next_level_name = self.gt['names'][ct+1]
            if next_level_name is None:
                break
            if self.xpr:
                fanout_name     = fanout_name.replace("•PR","")
                next_level_name = next_level_name.replace("•PR","")

            assert len(fanout_name) < 31
            ws_level = wb.create_sheet(fanout_name.replace("/"," "))
            sheet_title = f"{self.gt['name']}: fanout from {self.gt['names'][ct]} to {self.gt['names'][ct+1]}"
            ws_setup_level(ws_level,sheet_title)
            geounits = self.get_geounits(ct)
            logging.info(f"ct: {ct} len(geounits)={len(geounits)} t={time.time()-t0}")

            # Turn the geounits into a queue for rapid access
            geounits = deque(geounits)
            # Now find all of the fanout groups
            level_stats = []
            row         = 3
            state       = None
            print_count = 0

            # Crunch through the returned geounits, dividing between each state, and summarizing on each row.
            all_fanout_populations = []
            while geounits:
                res = {}
                fanout_populations = []
                reporting_prefix = geounits[0]['reporting_prefix']
                while geounits and (reporting_prefix == geounits[0]['reporting_prefix']):
                    d0 = geounits.popleft()
                    if print_count < 10:
                        logging.info(dict(d0))
                        print_count += 1
                    fanout_populations.append(d0['population'])
                    all_fanout_populations.append(d0['population'])

                if state!=d0['state']:
                    # New state!
                    state           = d0['state']
                    # If page is below P1, add a space between each state
                    if ct>1:    
                        row += 1

                res['fanout_populations'] = fanout_populations
                res['fanout_count']       = len(fanout_populations)
                level_stats.append(res)

                # Populate the per-level information with the last d0 data and info for this res
                if ct==0:
                    column1_label = 'US'
                else:
                    column1_label = constants.STATE_TO_STUSAB[d0['state']]
                    if d0['reporting_prefix'][3:4]=='A':
                        column1_label += '-AIANHH'
                ws_level.cell(row=row,column=1).value = column1_label
                ws_level.cell(row=row,column=2).value = "_"+d0['reporting_prefix']
                if args.names:
                    ws_level.cell(row=row,column=3).value = gs.county_name(state, county)
                ws_level.cell( row=row,column=4).value = len(fanout_populations)
                ws_level.cell( row=row,column=5).value = sum(fanout_populations)
                ws_level.cell( row=row,column=6).value = int(statistics.mean(fanout_populations))
                for cellrow in ws_level.iter_rows(min_row=row, max_row=row,min_col=3,max_col=6):
                    for cell in cellrow:
                        cell.border = right_border
                ws_level.cell( row=row,column=6).border = right_thick_border

                # add commas to pop_tot and pop_avg
                for cellrow in ws_level.iter_rows(min_row=row, max_row=row,min_col=5,max_col=6):
                    for cell in cellrow:
                        cell.number_format = '#,##0'
                # layout and format deciles
                for cellrow in ws_level.iter_rows(min_row=row, max_row=row,min_col=7,max_col=17):
                    for(cell,value) in zip(cellrow, deciles(fanout_populations)):
                        cell.value = value
                        cell.number_format = '#,##0'
                ws_level.cell( row=row,column=17).border = right_thick_border
                row += 1

                if row % 10_000==0:
                    logging.info("written [%s] row %s",fanout_name, row)

            logging.info("total rows=%s t=%s",row,time.time()-t0)
            # Finalize the Sheet
            ws_level.freeze_panes    = 'A3'
            ws_level.auto_filter.ref = f'A2:Q{row-1}'
            wb.format_rows(ws_level, min_row=3, max_row=row-1, min_col=1, max_col=17, column=1,
                           stripe = (ct>1),
                           fills=[STATE1_FILL, STATE2_FILL], value_fills = {'PR':PR_FILL})

            # Now fill out the Overview Sheet
            ws_overview.cell(row=overview_row, column=1).value = fanout_name
            ws_overview.cell(row=overview_row, column=2).value = len(level_stats)
            ws_overview.cell(row=overview_row, column=3).value = next_level_name
            ws_overview.cell(row=overview_row, column=4).value = sum([res['fanout_count'] for res in level_stats])
            logging.info("fanout_name=%s next_level_name=%s",fanout_name,next_level_name)

            fanouts = [res['fanout_count'] for res in level_stats]
            fanout_deciles = deciles(fanouts)
            for cellrow in ws_overview.iter_rows(min_row = overview_row, max_row=overview_row, min_col=5, max_col = 15):
                for (cell,value) in zip(cellrow,fanout_deciles):
                    cell.value = value
                    cell.fill = YELLOW_FILL
                    cell.border = thin_border

            fanout_population_deciles = deciles(all_fanout_populations)
            for cellrow in ws_overview.iter_rows(min_row = overview_row, max_row=overview_row, min_col=16, max_col = 26):
                for (cell,value) in zip(cellrow,fanout_population_deciles):
                    cell.value = value
                    cell.number_format = "#,##0"
                    cell.fill = PINK_FILL
                    cell.border = thin_border
            overview_row += 1
            # Save this level


            fname = f"{fnamebase} {ct}.xlsx"
            logging.info("Saving %s",fname)
            with ctools.timer.Timer(notifier=logging.info):
                wb.save(fname)
            if ct+1==args.levels:
                break
        ws_add_notes(ws_overview,overview_row+2,"geotree_notes.md")
        ws_add_metadata(wb)
        fname = f"{fnamebase}.xlsx"
        logging.info("Saving %s",fname)
        with ctools.timer.Timer(notifier=logging.info):
            wb.save(fname)
        subprocess.call(['open',fname])


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Ingest the PL94 block-level population counts',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--db", help="Specify database location", default=pl94_dbload.DBFILE)
    parser.add_argument("--create", action='store_true', help='create the schema')
    parser.add_argument("--delete", action='store_true', help='delete the schema')
    parser.add_argument("--dump",   action='store_true', help='print the blocks')
    parser.add_argument("--scheme" , help='specify partitioning scheme')
    parser.add_argument("--report", action='store_true', help="Create a report")
    parser.add_argument("--names", action='store_true', help='display names')
    parser.add_argument("--levels", type=int, help="how many levels")
    parser.add_argument("--xpr",     action='store_true', help='remove PR from reports')
    ctools.clogging.add_argument(parser)
    args = parser.parse_args()
    args.name = 'table' + args.scheme[1]

    ctools.clogging.setup(level=args.loglevel)
    gc.enable()

    db   = ctools.dbfile.DBSqlite3(args.db,dicts=True,debug=False)
    db.set_cache_bytes(4*1024*1024*1024)

    gt = GeoTree(db,args.name,args.scheme,args.xpr)

    # open database and give me a big cache
    if args.delete:
        db.execute(f"DROP TABLE IF EXISTS {args.name}",debug=True)
        db.execute(f"DROP INDEX IF EXISTS {args.name}_logrecno",debug=True)
        db.execute(f"DROP INDEX IF EXISTS {args.name}_p",debug=True)

    if args.create:
        gt.create()
                         
    if args.dump:
        gt.dump()

    if args.report:
        gt.report()
